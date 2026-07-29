import logging
import re
from decimal import Decimal, InvalidOperation

import openpyxl

from app.catalogo.types import ComponenteImportado

logger = logging.getLogger(__name__)

FAMILIAS_TERMOMAGNETICO = {
    "Interruptores Termomagnéticos",
    "Interruptores Termomagnéticos - Con posibilidad de utilizar accesorios",
    "Interruptores Termomagnéticos - Sin posibilidad de utilizar accesorios",
    "Interruptores automáticos en caja moldeada",
}
FAMILIAS_DIFERENCIAL = {
    "Interruptores termomagnéticos con protección diferencial",
    "Interruptores Diferenciales",
    "Interruptores Diferenciales - Sin posibilidad de utilizar accesorios",
    "Interruptores Diferenciales - Con posibilidad de utilizar accesorios",
    "Detector de fallas de arco con proteccion Diferencial (AFDD+RCD)",
}

_POLOS_MAP = {"uni": 1, "bi": 2, "tri": 3, "tetra": 4}
_POLOS_DESCRIPCION_RE = re.compile(r"\b(uni|bi|tri|tetra)polar", re.IGNORECASE)
_POLOS_CATEGORIA_RE = re.compile(r"\b(uni|bi|tri|tetra)polar(es)?\b", re.IGNORECASE)
_CORRIENTE_RE = re.compile(r"In\s*=?\s*(\d+(?:[.,]\d+)?)")
_CAPACIDAD_DESCRIPCION_RE = re.compile(r"Ic[nu]\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*kA", re.IGNORECASE)
_CAPACIDAD_CATEGORIA_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*kA", re.IGNORECASE)


_SENSIBILIDAD_DESCRIPCION_RE = re.compile(
    r"(?:Sens(?:ibilidad)?|I[Δ∆]?)\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*(m?A)",
    re.IGNORECASE,
)


def _texto_a_decimal(texto: str) -> Decimal:
    return Decimal(texto.replace(",", "."))


def _extraer_polos(categoria_path: list[str], descripcion: str) -> int | None:
    match = _POLOS_DESCRIPCION_RE.search(descripcion or "")
    if match:
        return _POLOS_MAP[match.group(1).lower()]
    for nivel in categoria_path:
        match = _POLOS_CATEGORIA_RE.search(nivel)
        if match:
            return _POLOS_MAP[match.group(1).lower()]
    return None


def _extraer_corriente_nominal(descripcion: str) -> Decimal | None:
    match = _CORRIENTE_RE.search(descripcion or "")
    if not match:
        return None
    return _texto_a_decimal(match.group(1))


def _extraer_capacidad_corte_termomagnetico(descripcion: str) -> Decimal | None:
    valores = _CAPACIDAD_DESCRIPCION_RE.findall(descripcion or "")
    if not valores:
        return None
    return min(_texto_a_decimal(v) for v in valores)


def _extraer_capacidad_corte_combo(categoria_path: list[str]) -> Decimal | None:
    if len(categoria_path) < 2:
        return None
    match = _CAPACIDAD_CATEGORIA_RE.search(categoria_path[1])
    if not match:
        return None
    return _texto_a_decimal(match.group(1))


def _extraer_sensibilidad_ma(categoria_path: list[str], descripcion: str) -> int | None:
    texto_busqueda = (descripcion or "") + " " + " ".join(categoria_path or [])
    match = _SENSIBILIDAD_DESCRIPCION_RE.search(texto_busqueda)
    if match:
        val_str = match.group(1).replace(",", ".")
        unidad = match.group(2).lower()
        val = float(val_str)
        if unidad == "a":
            return int(round(val * 1000))
        return int(round(val))

    match_ma = re.search(r"\b(\d+)\s*mA\b", texto_busqueda, re.IGNORECASE)
    if match_ma:
        return int(match_ma.group(1))

    return None


def _extraer_accesorios(categoria_path: list[str], descripcion: str) -> bool | None:
    texto_completo = " ".join(categoria_path or []) + " " + (descripcion or "")
    if re.search(r"sin\s+posibilidad", texto_completo, re.IGNORECASE):
        return False
    if re.search(r"con\s+posibilidad", texto_completo, re.IGNORECASE):
        return True
    return None


def _extraer_atributos(categoria_path: list[str], descripcion: str) -> dict | None:
    if not categoria_path:
        return None
    raiz = categoria_path[0]

    if raiz in FAMILIAS_TERMOMAGNETICO:
        tipo = "seccional_termomagnetico"
        polos = _extraer_polos(categoria_path, descripcion)
        corriente = _extraer_corriente_nominal(descripcion)
        capacidad = _extraer_capacidad_corte_termomagnetico(descripcion)
    elif raiz in FAMILIAS_DIFERENCIAL:
        tipo = "seccional_diferencial"
        polos = _extraer_polos(categoria_path, descripcion)
        corriente = _extraer_corriente_nominal(descripcion)
        capacidad = (
            _extraer_capacidad_corte_combo(categoria_path)
            or _extraer_capacidad_corte_termomagnetico(descripcion)
            or Decimal("10.0")
        )
    else:
        return None

    if polos is None or corriente is None or capacidad is None:
        return None

    res = {
        "tipo": tipo,
        "polos": polos,
        "corriente_nominal_a": float(corriente),
        "capacidad_corte_ka": float(capacidad),
    }

    sensibilidad = _extraer_sensibilidad_ma(categoria_path, descripcion)
    if sensibilidad is not None:
        res["sensibilidad_ma"] = sensibilidad

    accesorios = _extraer_accesorios(categoria_path, descripcion)
    if accesorios is not None:
        res["admite_accesorios"] = accesorios

    return res


def parse_abb_workbook(file_obj, archivo_origen: str) -> list[ComponenteImportado]:
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[_find_lista_de_precios_sheet(wb)]
    header_map = _read_header_map(ws, header_row=1)

    resultados: list[ComponenteImportado] = []
    path: list[tuple[tuple[float, bool], str]] = []

    codigo_col = _get_required_column(header_map, "Codigo SAP")
    comercial_col = _get_required_column(header_map, "Codigo Comercial")

    for row_idx in range(3, ws.max_row + 1):
        codigo_cell = ws.cell(row=row_idx, column=codigo_col)
        comercial_cell = ws.cell(row=row_idx, column=comercial_col)
        # Real en el Excel de ABB: algunas filas de sección tienen "Codigo SAP"
        # en blanco pero no None (un espacio u otro whitespace) -- tratarlas
        # igual que None evita que se cuelen como "componentes" fantasma con
        # código vacío.
        tiene_codigo = codigo_cell.value is not None and str(codigo_cell.value).strip()

        if not tiene_codigo and comercial_cell.value is not None:
            texto = str(comercial_cell.value).strip()
            if texto:
                firma = (comercial_cell.font.sz, bool(comercial_cell.font.bold))
                _update_path(path, firma, texto)
        elif tiene_codigo:
            resultados.append(_build_componente(ws, row_idx, header_map, path, archivo_origen))

    return resultados


def _find_lista_de_precios_sheet(wb) -> str:
    candidatos = [s for s in wb.sheetnames if s.strip().lower().startswith("lista de precios")]
    if len(candidatos) != 1:
        raise ValueError(f"Se esperaba exactamente una pestaña 'Lista de Precios...', se encontraron: {candidatos}")
    return candidatos[0]


def _read_header_map(ws, header_row: int) -> dict[str, int]:
    header_map = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value:
            header_map[str(value).strip()] = col
    return header_map


def _get_required_column(header_map: dict[str, int], col: str) -> int:
    if col not in header_map:
        raise ValueError(f"Columna requerida '{col}' no encontrada en la hoja")
    return header_map[col]


def _update_path(path: list[tuple[tuple, str]], firma: tuple, texto: str) -> None:
    for i, (existing_firma, _) in enumerate(path):
        if existing_firma == firma:
            del path[i:]
            break
    path.append((firma, texto))


def _decimal_or_none(value, row_idx: int) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        # Real-world price lists use text placeholders like "Consultar" ("price on
        # request") instead of a numeric value. Treat that the same as a blank
        # price cell rather than aborting the whole import over one bad row.
        logger.warning("Precio no numérico en fila %s: %r — se importa sin precio", row_idx, value)
        return None


def _build_componente(ws, row_idx, header_map, path, archivo_origen) -> ComponenteImportado:
    def cell(label):
        col = header_map.get(label)
        return ws.cell(row=row_idx, column=col).value if col else None

    comercial = cell("Codigo Comercial")
    categoria_path = [texto for _, texto in path]
    descripcion = str(cell("Descripcion") or "").strip()
    return ComponenteImportado(
        proveedor="ABB",
        codigo=str(cell("Codigo SAP")).strip(),
        codigo_comercial=str(comercial).strip() if comercial else None,
        categoria_path=categoria_path,
        descripcion=descripcion,
        unidad="Unidad",
        precio_lista=_decimal_or_none(cell("Precio de Lista USD"), row_idx),
        precio_neto=_decimal_or_none(cell("Precio NETO USD"), row_idx),
        atributos=_extraer_atributos(categoria_path, descripcion),
        archivo_origen=archivo_origen,
        fila_origen=row_idx,
    )
