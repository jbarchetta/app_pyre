import logging
from decimal import Decimal, InvalidOperation

import openpyxl

from app.catalogo.types import ComponenteImportado

logger = logging.getLogger(__name__)


def parse_otros_workbook(
    file_obj, archivo_origen: str, sheet_name: str = "Lista de Precios"
) -> list[ComponenteImportado]:
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[sheet_name]

    resultados: list[ComponenteImportado] = []
    categoria_raiz: str | None = None
    subfamilia: str | None = None
    columna_map: dict[str, int] | None = None

    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value

        if col_a is not None:
            categoria_raiz = str(col_a).strip()
            subfamilia = None
            columna_map = None
            continue

        if col_b is None:
            continue

        texto_b = str(col_b).strip()

        if texto_b == "Cod":
            columna_map = _read_row_labels(ws, row_idx)
            continue

        if columna_map is None:
            subfamilia = texto_b
            continue

        if _is_subfamilia_label_row(ws, row_idx, columna_map):
            subfamilia = texto_b
            columna_map = None
            continue

        resultados.append(_build_componente(ws, row_idx, columna_map, categoria_raiz, subfamilia, archivo_origen))

    return resultados


def _is_subfamilia_label_row(ws, row_idx: int, columna_map: dict[str, int]) -> bool:
    otras_columnas = [col for col in columna_map.values() if col != 2]
    return all(ws.cell(row=row_idx, column=col).value is None for col in otras_columnas)


def _read_row_labels(ws, row_idx: int) -> dict[str, int]:
    labels = {}
    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=row_idx, column=col).value
        if value:
            labels[str(value).strip()] = col
    return labels


def _decimal_or_none(value, row_idx: int) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        # Real-world price lists use placeholders like "#DIV/0!" (a broken Excel
        # formula) instead of a numeric value. Treat that the same as a blank
        # price cell rather than aborting the whole import over one bad row.
        logger.warning("Precio no numérico en fila %s: %r — se importa sin precio", row_idx, value)
        return None


def _build_componente(
    ws, row_idx, columna_map: dict[str, int], categoria_raiz: str | None, subfamilia: str | None, archivo_origen: str
) -> ComponenteImportado:
    def cell(label):
        col = columna_map.get(label)
        return ws.cell(row=row_idx, column=col).value if col else None

    categoria_path = ([categoria_raiz] if categoria_raiz else []) + ([subfamilia] if subfamilia else [])
    precio_lista = _decimal_or_none(cell("Precio Lista ((U$S)"), row_idx)
    precio_neto = _decimal_or_none(cell("Total U$S)") if cell("Total U$S)") is not None else cell("Total"), row_idx)

    return ComponenteImportado(
        proveedor="OTROS",
        codigo=str(cell("Cod") or "").strip(),
        codigo_comercial=None,
        categoria_path=categoria_path,
        descripcion=str(cell("Descripcion") or "").strip(),
        unidad=str(cell("Unidad") or "Unidad").strip(),
        precio_lista=precio_lista,
        precio_neto=precio_neto,
        archivo_origen=archivo_origen,
        fila_origen=row_idx,
    )
