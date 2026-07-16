import io
from decimal import Decimal

import openpyxl
import pytest

from app.catalogo.parser_otros import parse_otros_workbook


def _build_sample_workbook() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    ws.cell(row=4, column=1, value="Accesorios Tableros")
    ws.cell(row=6, column=2, value="Cable Canal Ranurado")
    header_cols = {2: "Cod", 3: "Unidad", 4: "Descripcion", 7: "Precio Lista ((U$S)", 10: "Total U$S)"}
    for col, label in header_cols.items():
        ws.cell(row=8, column=col, value=label)
    ws.cell(row=9, column=2, value="A00000")
    ws.cell(row=9, column=3, value="Un.")
    ws.cell(row=9, column=4, value="Cable Canal Ranurado 15x15")
    ws.cell(row=9, column=10, value=4.78)

    ws.cell(row=12, column=1, value="Gabinetes")
    ws.cell(row=13, column=2, value="SELECCION DE GABINETES A MEDIDA")
    ws.cell(row=14, column=4, value="NOLLMED")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_parses_data_rows_with_local_header_and_breadcrumb():
    resultados = parse_otros_workbook(_build_sample_workbook(), archivo_origen="test.xlsx")

    assert len(resultados) == 1
    item = resultados[0]
    assert item.codigo == "A00000"
    assert item.categoria_path == ["Accesorios Tableros", "Cable Canal Ranurado"]
    assert item.precio_neto == Decimal("4.78")
    assert item.precio_lista is None
    assert item.proveedor == "OTROS"
    assert item.unidad == "Un."
    assert item.fila_origen == 9


def test_section_without_cod_header_yields_no_rows():
    resultados = parse_otros_workbook(_build_sample_workbook(), archivo_origen="test.xlsx")

    gabinetes = [r for r in resultados if r.categoria_path[0] == "Gabinetes"]
    assert gabinetes == []


def _build_multi_subfamilia_workbook() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    ws.cell(row=4, column=1, value="Accesorios Tableros")

    ws.cell(row=6, column=2, value="Cable Canal Ranurado")
    header_cols_1 = {2: "Cod", 3: "Unidad", 4: "Descripcion", 7: "Precio Lista ((U$S)", 10: "Total U$S)"}
    for col, label in header_cols_1.items():
        ws.cell(row=8, column=col, value=label)
    ws.cell(row=9, column=2, value="A00000")
    ws.cell(row=9, column=3, value="Un.")
    ws.cell(row=9, column=4, value="Cable Canal Ranurado 15x15")
    ws.cell(row=9, column=10, value=4.78)

    ws.cell(row=11, column=2, value="Terminales")
    header_cols_2 = {2: "Cod", 3: "Unidad", 4: "Descripcion", 10: "Total U$S)"}
    for col, label in header_cols_2.items():
        ws.cell(row=13, column=col, value=label)
    ws.cell(row=14, column=2, value="B00000")
    ws.cell(row=14, column=3, value="Un.")
    ws.cell(row=14, column=4, value="Terminal 5")
    ws.cell(row=14, column=10, value=1.23)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_second_subfamilia_under_same_category_gets_own_header_and_breadcrumb():
    resultados = parse_otros_workbook(_build_multi_subfamilia_workbook(), archivo_origen="test.xlsx")

    assert len(resultados) == 2

    primero = next(r for r in resultados if r.codigo == "A00000")
    assert primero.categoria_path == ["Accesorios Tableros", "Cable Canal Ranurado"]
    assert primero.precio_neto == Decimal("4.78")

    segundo = next(r for r in resultados if r.codigo == "B00000")
    assert segundo.categoria_path == ["Accesorios Tableros", "Terminales"]
    assert segundo.precio_neto == Decimal("1.23")


def test_raises_clear_value_error_on_non_numeric_price():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    ws.cell(row=4, column=1, value="Accesorios Tableros")
    ws.cell(row=6, column=2, value="Cable Canal Ranurado")
    header_cols = {2: "Cod", 3: "Unidad", 4: "Descripcion", 7: "Precio Lista ((U$S)", 10: "Total U$S)"}
    for col, label in header_cols.items():
        ws.cell(row=8, column=col, value=label)
    ws.cell(row=9, column=2, value="A00000")
    ws.cell(row=9, column=3, value="Un.")
    ws.cell(row=9, column=4, value="Cable Canal Ranurado 15x15")
    ws.cell(row=9, column=10, value="Consultar")  # non-numeric placeholder in a price column

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    with pytest.raises(ValueError):
        parse_otros_workbook(buffer, archivo_origen="test.xlsx")
