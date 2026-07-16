import io

from openpyxl import Workbook
from openpyxl.styles import Font

from app.scripts.create_user import create_user


def _sample_abb_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"
    headers = [
        "Codigo SAP", "Codigo Comercial", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    cat = ws.cell(row=3, column=2, value="Interruptores Termomagneticos")
    cat.font = Font(size=14, bold=False)
    fam = ws.cell(row=4, column=2, value="SH200 L")
    fam.font = Font(size=14, bold=True)
    curva = ws.cell(row=6, column=2, value="Curva C - Icn: 4,5kA")
    curva.font = Font(size=10, bold=True)
    polos = ws.cell(row=7, column=2, value="Unipolares")
    polos.font = Font(size=12, bold=False)

    ws.cell(row=8, column=1, value="COD-U2")
    ws.cell(row=8, column=2, value="SH201-C2")
    ws.cell(row=8, column=9, value=15.4)
    ws.cell(row=8, column=10, value=7.8)
    ws.cell(row=8, column=16, value="Interruptor unipolar In 2A")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _sample_otros_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    ws.cell(row=4, column=1, value="Accesorios Tableros")
    ws.cell(row=6, column=2, value="Cable Canal Ranurado")
    header_cols = {2: "Cod", 3: "Unidad", 4: "Descripcion", 7: "Precio Lista ((U$S)", 10: "Total U$S)"}
    for col, label in header_cols.items():
        ws.cell(row=8, column=col, value=label)
    ws.cell(row=9, column=2, value="A00000")
    ws.cell(row=9, column=3, value="Un.")
    ws.cell(row=9, column=4, value="Cable Canal Ranurado 15x15")
    ws.cell(row=9, column=10, value=4.78)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _xlsx_file_tuple(filename: str = "abb.xlsx"):
    return (
        filename,
        _sample_abb_bytes(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _otros_xlsx_file_tuple(filename: str = "otros.xlsx"):
    return (
        filename,
        _sample_otros_bytes(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_import_requires_authentication(client):
    response = client.post(
        "/catalogo/importar",
        data={"proveedor": "abb"},
        files={"archivo": _xlsx_file_tuple()},
    )

    assert response.status_code == 401


def test_import_abb_catalog_end_to_end(client, db_session):
    create_user("import.endpoint.test@pyre.com", "Importador", "clave-segura-123", "analista", db=db_session)
    client.post("/auth/login", json={"email": "import.endpoint.test@pyre.com", "password": "clave-segura-123"})

    response = client.post(
        "/catalogo/importar",
        data={"proveedor": "abb"},
        files={"archivo": _xlsx_file_tuple()},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["nuevos"] == 1
    assert body["total_filas"] == 1


def test_import_rejects_unknown_proveedor(client, db_session):
    create_user("import.unknown.test@pyre.com", "Importador", "clave-segura-123", "analista", db=db_session)
    client.post("/auth/login", json={"email": "import.unknown.test@pyre.com", "password": "clave-segura-123"})

    response = client.post(
        "/catalogo/importar",
        data={"proveedor": "no-existe"},
        files={"archivo": _xlsx_file_tuple()},
    )

    assert response.status_code == 400


def test_import_rejects_corrupted_file(client, db_session):
    create_user("import.corrupt.test@pyre.com", "Importador", "clave-segura-123", "analista", db=db_session)
    client.post("/auth/login", json={"email": "import.corrupt.test@pyre.com", "password": "clave-segura-123"})

    response = client.post(
        "/catalogo/importar",
        data={"proveedor": "abb"},
        files={
            "archivo": (
                "bad.xlsx",
                b"not a real xlsx file",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400


def test_import_otros_catalog_end_to_end(client, db_session):
    create_user("import.otros.test@pyre.com", "Importador", "clave-segura-123", "analista", db=db_session)
    client.post("/auth/login", json={"email": "import.otros.test@pyre.com", "password": "clave-segura-123"})

    response = client.post(
        "/catalogo/importar",
        data={"proveedor": "otros"},
        files={"archivo": _otros_xlsx_file_tuple()},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["nuevos"] == 1
    assert body["total_filas"] == 1
