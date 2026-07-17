import io
from decimal import Decimal

import openpyxl
import pytest
from openpyxl.styles import Font

from app.catalogo.parser_abb import parse_abb_workbook


def _build_sample_workbook() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"

    headers = [
        "Codigo SAP", "Codigo Comercial", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    def header_row(row, text, size, bold):
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = Font(size=size, bold=bold)

    def data_row(row, codigo, comercial, precio_lista, precio_neto, descripcion):
        ws.cell(row=row, column=1, value=codigo)
        ws.cell(row=row, column=2, value=comercial)
        ws.cell(row=row, column=9, value=precio_lista)
        ws.cell(row=row, column=10, value=precio_neto)
        ws.cell(row=row, column=16, value=descripcion)

    header_row(3, "Interruptores Termomagneticos", 14, False)
    header_row(4, "SH200 L", 14, True)
    header_row(6, "Curva C - Icn: 4,5kA (IEC 60898)", 10, True)
    header_row(7, "Unipolares", 12, False)
    data_row(8, "COD-U2", "SH201-C2", 15.4, 7.8, "Interruptor unipolar In 2A")
    data_row(9, "COD-U4", "SH201-C4", 15.4, 7.8, "Interruptor unipolar In 4A")
    header_row(10, "Bipolares", 12, False)
    data_row(11, "COD-B2", "SH202-C2", 20.1, 10.2, "Interruptor bipolar In 2A")
    header_row(12, "Interruptores Diferenciales", 14, False)
    header_row(13, "F200", 14, True)
    header_row(14, "30mA", 10, True)
    header_row(15, "Bipolares", 12, False)
    data_row(16, "COD-DIF-B", "F202-30", 50.0, 25.0, "Diferencial bipolar 30mA")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_parses_sections_with_correct_breadcrumbs_and_prices():
    resultados = parse_abb_workbook(_build_sample_workbook(), archivo_origen="test.xlsx")

    assert len(resultados) == 4

    unipolar_2a = next(r for r in resultados if r.codigo == "COD-U2")
    assert unipolar_2a.categoria_path == [
        "Interruptores Termomagneticos", "SH200 L", "Curva C - Icn: 4,5kA (IEC 60898)", "Unipolares",
    ]
    assert unipolar_2a.precio_lista == Decimal("15.4")
    assert unipolar_2a.precio_neto == Decimal("7.8")
    assert unipolar_2a.proveedor == "ABB"
    assert unipolar_2a.codigo_comercial == "SH201-C2"
    assert unipolar_2a.descripcion == "Interruptor unipolar In 2A"
    assert unipolar_2a.fila_origen == 8

    bipolar_2a = next(r for r in resultados if r.codigo == "COD-B2")
    # same family/curve breadcrumb, "Bipolares" replaces "Unipolares" at that level
    assert bipolar_2a.categoria_path == [
        "Interruptores Termomagneticos", "SH200 L", "Curva C - Icn: 4,5kA (IEC 60898)", "Bipolares",
    ]

    diferencial = next(r for r in resultados if r.codigo == "COD-DIF-B")
    # a brand new top-level category resets the whole breadcrumb
    assert diferencial.categoria_path == ["Interruptores Diferenciales", "F200", "30mA", "Bipolares"]


def test_raises_when_no_lista_de_precios_sheet_found():
    wb = openpyxl.Workbook()
    wb.active.title = "Otra Hoja"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    with pytest.raises(ValueError):
        parse_abb_workbook(buffer, archivo_origen="test.xlsx")


def test_non_numeric_price_placeholder_parses_as_none_instead_of_raising():
    # 'Consultar' ("price on request") shows up in ~0.5% of rows in the real ABB
    # price list -- it's a normal commercial value, not corrupted data, so the row
    # should still import with that price left as None rather than aborting the
    # whole file.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"

    headers = [
        "Codigo SAP", "Codigo Comercial", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    ws.cell(row=3, column=1, value="COD-BAD")
    ws.cell(row=3, column=2, value="X-1")
    ws.cell(row=3, column=9, value="Consultar")  # non-numeric placeholder in a price column
    ws.cell(row=3, column=10, value=10.0)
    ws.cell(row=3, column=16, value="Componente con precio inválido")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resultados = parse_abb_workbook(buffer, archivo_origen="test.xlsx")

    assert len(resultados) == 1
    componente = resultados[0]
    assert componente.codigo == "COD-BAD"
    assert componente.precio_lista is None
    assert componente.precio_neto == Decimal("10.0")


def test_raises_clear_value_error_when_required_header_missing():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"

    # "Codigo Comercial" header (column 2) intentionally omitted/renamed.
    headers = [
        "Codigo SAP", "Otra Columna", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    ws.cell(row=3, column=1, value="COD-1")
    ws.cell(row=3, column=2, value="X-1")
    ws.cell(row=3, column=9, value=15.4)
    ws.cell(row=3, column=10, value=7.8)
    ws.cell(row=3, column=16, value="Componente")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    with pytest.raises(ValueError):
        parse_abb_workbook(buffer, archivo_origen="test.xlsx")


def test_skips_blank_row_mixed_in_among_data_rows():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"

    headers = [
        "Codigo SAP", "Codigo Comercial", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    def data_row(row, codigo, comercial, precio_lista, precio_neto, descripcion):
        ws.cell(row=row, column=1, value=codigo)
        ws.cell(row=row, column=2, value=comercial)
        ws.cell(row=row, column=9, value=precio_lista)
        ws.cell(row=row, column=10, value=precio_neto)
        ws.cell(row=row, column=16, value=descripcion)

    data_row(3, "COD-1", "X-1", 15.4, 7.8, "Primer componente")
    # Row 4 is entirely blank: both Codigo SAP and Codigo Comercial are None.
    data_row(5, "COD-2", "X-2", 20.1, 10.2, "Segundo componente")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resultados = parse_abb_workbook(buffer, archivo_origen="test.xlsx")

    assert len(resultados) == 2
    assert {r.codigo for r in resultados} == {"COD-1", "COD-2"}
    assert all(r.fila_origen != 4 for r in resultados)


def test_parse_abb_workbook_populates_atributos_for_in_scope_rows():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"

    headers = [
        "Codigo SAP", "Codigo Comercial", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    def header_row(row, text, size, bold):
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = Font(size=size, bold=bold)

    def data_row(row, codigo, comercial, precio_lista, precio_neto, descripcion):
        ws.cell(row=row, column=1, value=codigo)
        ws.cell(row=row, column=2, value=comercial)
        ws.cell(row=row, column=9, value=precio_lista)
        ws.cell(row=row, column=10, value=precio_neto)
        ws.cell(row=row, column=16, value=descripcion)

    header_row(3, "Interruptores Termomagnéticos", 14, False)
    header_row(4, "SH200 L", 14, True)
    header_row(6, "Curva C - Icn: 4,5kA (IEC 60898)", 10, True)
    header_row(7, "Unipolares", 12, False)
    data_row(8, "COD-U2", "SH201-C2", 15.4, 7.8, "Interruptor termomagnético unipolar In 2A Icn = 4,5kA @ IEC60898 Curva C")
    header_row(9, "Interruptores Diferenciales", 14, False)
    header_row(10, "F200", 14, True)
    header_row(11, "30mA", 10, True)
    header_row(12, "Bipolares", 12, False)
    data_row(13, "COD-DIF", "F202-30", 50.0, 25.0, "Interruptor diferencial bipolar In 16. Sens = 10 mA")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resultados = parse_abb_workbook(buffer, archivo_origen="test.xlsx")

    termomagnetico = next(r for r in resultados if r.codigo == "COD-U2")
    assert termomagnetico.atributos == {
        "tipo": "seccional_termomagnetico",
        "polos": 1,
        "corriente_nominal_a": 2.0,
        "capacidad_corte_ka": 4.5,
    }

    diferencial = next(r for r in resultados if r.codigo == "COD-DIF")
    assert diferencial.atributos is None


def test_fila_con_codigo_string_vacio_se_trata_como_breadcrumb_no_como_componente():
    # Bug real encontrado en el Excel de ABB (~1.185 de 10.247 filas): algunas
    # filas de sección tienen la celda "Codigo SAP" como string vacío (""), no
    # None -- antes se colaban como "componentes" fantasma con código vacío (y
    # a veces con el texto de una nota al pie como codigo_comercial, que
    # revienta el límite de varchar(100) de esa columna al importar).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios 202607"

    headers = [
        "Codigo SAP", "Codigo Comercial", None, None, None, None, None, None,
        "Precio de Lista USD", "Precio NETO USD", None, None, None, None, None, "Descripcion",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)

    def header_row(row, text, size, bold):
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = Font(size=size, bold=bold)

    header_row(3, "Interruptores Termomagneticos", 14, False)
    # " " (espacio) en vez de "" -- openpyxl normaliza "" a None al guardar el
    # archivo, lo que no reproduce el caso real (una celda con contenido en
    # blanco pero no None). Un espacio sobrevive el guardado/lectura igual que
    # el caso real del Excel de ABB.
    ws.cell(row=4, column=1, value=" ")
    header_row(4, "SH200 L", 14, True)

    ws.cell(row=5, column=1, value="COD-U2")
    ws.cell(row=5, column=2, value="SH201-C2")
    ws.cell(row=5, column=9, value=15.4)
    ws.cell(row=5, column=10, value=7.8)
    ws.cell(row=5, column=16, value="Interruptor unipolar In 2A")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resultados = parse_abb_workbook(buffer, archivo_origen="test.xlsx")

    assert len(resultados) == 1
    assert resultados[0].codigo == "COD-U2"
    assert resultados[0].categoria_path == ["Interruptores Termomagneticos", "SH200 L"]
